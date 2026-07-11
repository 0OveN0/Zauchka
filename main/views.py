from datetime import datetime, timedelta
import os
import mimetypes
from types import SimpleNamespace

import openpyxl
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from django.db.models import Q
from django.db.utils import OperationalError, ProgrammingError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.conf import settings
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.http import require_POST

from .models import (
    BannedWord,
    Comment,
    CustomTheme,
    FavoriteItem,
    Olympiad,
    Profile,
    SiteSetting,
    SupportMessage,
    SupportThread,
    ThemePreset,
)


MAX_SUPPORT_FILE_SIZE = 5 * 1024 * 1024
ONLINE_TIMEOUT = timedelta(minutes=2)

def is_site_admin(user):
    return bool(user.is_authenticated and (user.is_superuser or user.is_staff or user.username == 'admin'))


DEFAULT_THEME_PRESETS = [
    {
        'key': 'telegram-light', 'name': 'Голубая Telegram', 'is_builtin': True, 'sort_order': 10,
        'background_color': '#eaf6ff', 'card_color': '#ffffff', 'card_transparency': 0.82, 'card_blur': 22,
        'card_style': 'telegram', 'text_color': '#17212b', 'muted_color': '#5f7285', 'primary_color': '#229ED9',
        'background_size': 'cover', 'overlay_effect': 'none', 'overlay_opacity': 0.45,
    },
    {
        'key': 'black-orange', 'name': 'Black Orange', 'is_builtin': True, 'sort_order': 20,
        'background_color': '#0b0d10', 'card_color': '#171a20', 'card_transparency': 0.72, 'card_blur': 24,
        'card_style': 'dark_glass', 'text_color': '#fff7ed', 'muted_color': '#a8a29e', 'primary_color': '#f97316', 'danger_color': '#fb7185',
        'card_shadow_opacity': 0.28, 'overlay_effect': 'none', 'overlay_opacity': 0.55,
    },
    {
        'key': 'aqua-ios', 'name': 'Aqua iOS', 'is_builtin': True, 'sort_order': 30,
        'background_color': '#dff7ff', 'card_color': '#f8fdff', 'card_transparency': 0.62, 'card_blur': 34,
        'card_style': 'frosted', 'text_color': '#0f2537', 'muted_color': '#507086', 'primary_color': '#0ea5e9',
        'background_size': 'cover', 'overlay_effect': 'rain', 'overlay_opacity': 0.18,
    },
    {
        'key': 'pink-soft', 'name': 'Pink Soft', 'is_builtin': True, 'sort_order': 40,
        'background_color': '#fff1f7', 'card_color': '#ffffff', 'card_transparency': 0.78, 'card_blur': 24,
        'card_style': 'instagram', 'text_color': '#331827', 'muted_color': '#8b5f72', 'primary_color': '#ec4899', 'danger_color': '#e11d48',
        'overlay_effect': 'none', 'overlay_opacity': 0.28,
    },
    {
        'key': 'business', 'name': 'Business Metal', 'is_builtin': True, 'sort_order': 50,
        'background_color': '#cfd5dd', 'card_color': '#f3f6f9', 'card_transparency': 0.54, 'card_blur': 32,
        'card_style': 'metallic', 'text_color': '#17202a', 'muted_color': '#667085', 'primary_color': '#475569',
        'reflection_strength': 0.32, 'card_shadow_opacity': 0.20, 'overlay_effect': 'none', 'overlay_opacity': 0.32,
    },
    {
        'key': 'new-year', 'name': 'Новый год', 'is_builtin': True, 'sort_order': 60,
        'background_color': '#e9f8ff', 'card_color': '#ffffff', 'card_transparency': 0.70, 'card_blur': 24,
        'card_style': 'frosted', 'text_color': '#102a43', 'muted_color': '#5b7282', 'primary_color': '#0f9f8f', 'danger_color': '#dc2626',
        'overlay_effect': 'snow', 'overlay_opacity': 0.55,
    },
    {
        'key': 'custom', 'name': 'Своя тема', 'is_builtin': True, 'sort_order': 90,
        'background_color': '#f7fbff', 'card_color': '#ffffff', 'card_transparency': 0.86, 'card_blur': 18,
        'card_style': 'frosted', 'text_color': '#17212b', 'muted_color': '#657786', 'primary_color': '#229ED9',
        'overlay_effect': 'none', 'overlay_opacity': 0.35,
    },
]

LEGACY_THEME_KEYS = {item['key'] for item in DEFAULT_THEME_PRESETS}
ALLOWED_BACKGROUND_SIZES = {'contain', 'cover', '100% auto', 'auto 100%', 'auto'}
ALLOWED_BACKGROUND_POSITIONS = {'center center', 'center top', 'center bottom', 'left center', 'right center'}

RUSSIAN_REGIONS = [
    'Республика Адыгея', 'Республика Алтай', 'Республика Башкортостан', 'Республика Бурятия',
    'Республика Дагестан', 'Республика Ингушетия', 'Кабардино-Балкарская Республика',
    'Республика Калмыкия', 'Карачаево-Черкесская Республика', 'Республика Карелия',
    'Республика Коми', 'Республика Крым', 'Республика Марий Эл', 'Республика Мордовия',
    'Республика Саха (Якутия)', 'Республика Северная Осетия — Алания', 'Республика Татарстан',
    'Республика Тыва', 'Удмуртская Республика', 'Республика Хакасия', 'Чеченская Республика',
    'Чувашская Республика', 'Алтайский край', 'Забайкальский край', 'Камчатский край',
    'Краснодарский край', 'Красноярский край', 'Пермский край', 'Приморский край',
    'Ставропольский край', 'Хабаровский край', 'Амурская область', 'Архангельская область',
    'Астраханская область', 'Белгородская область', 'Брянская область', 'Владимирская область',
    'Волгоградская область', 'Вологодская область', 'Воронежская область', 'Запорожская область',
    'Ивановская область', 'Иркутская область', 'Калининградская область', 'Калужская область',
    'Кемеровская область — Кузбасс', 'Кировская область', 'Костромская область', 'Курганская область',
    'Курская область', 'Ленинградская область', 'Липецкая область', 'Магаданская область',
    'Московская область', 'Мурманская область', 'Нижегородская область', 'Новгородская область',
    'Новосибирская область', 'Омская область', 'Оренбургская область', 'Орловская область',
    'Пензенская область', 'Псковская область', 'Ростовская область', 'Рязанская область',
    'Самарская область', 'Саратовская область', 'Сахалинская область', 'Свердловская область',
    'Смоленская область', 'Тамбовская область', 'Тверская область', 'Томская область',
    'Тульская область', 'Тюменская область', 'Ульяновская область', 'Херсонская область',
    'Челябинская область', 'Ярославская область', 'Москва', 'Санкт-Петербург', 'Севастополь',
    'Еврейская автономная область', 'Ненецкий автономный округ', 'Ханты-Мансийский автономный округ — Югра',
    'Чукотский автономный округ', 'Ямало-Ненецкий автономный округ'
]

DEFAULT_RULES_CONTENT = '''<p><strong>Дата последнего обновления:</strong> 29.06.2026</p>

<h2>1. Общие положения</h2>
<p>Настоящее Пользовательское соглашение регулирует использование сайта Zauchka.RU.</p>

<h2>2. Регистрация и данные пользователя</h2>
<p>При регистрации пользователь обязуется указывать корректные данные и соблюдать правила общения.</p>

<h2>3. Поведение пользователей</h2>
<p>Запрещены спам, оскорбления, реклама и публикация материалов, нарушающих закон.</p>'''

DEFAULT_COOKIES_CONTENT = '''<p><strong>Дата последнего обновления:</strong> 29.06.2026</p>

<h2>1. Что такое cookie?</h2>
<p>Cookie помогают сайту сохранять авторизацию, выбранную тему и технические настройки интерфейса.</p>'''


def get_setting(key, default=''):
    obj, _ = SiteSetting.objects.get_or_create(key=key, defaults={'value': default})
    return obj.value


def set_setting(key, value):
    obj, _ = SiteSetting.objects.get_or_create(key=key, defaults={'value': value})
    obj.value = value
    obj.save()
    return obj


def clamp_int(value, default=0, minimum=0, maximum=10):
    try:
        number = int(value)
    except (TypeError, ValueError):
        return default
    return max(minimum, min(maximum, number))


def clamp_float(value, default=0.35, minimum=0.0, maximum=1.0):
    try:
        if isinstance(value, str):
            value = value.replace(',', '.').strip()
        number = float(value)
    except (TypeError, ValueError):
        return default
    return max(minimum, min(maximum, number))


def money_float(value, default=0.25, minimum=0.0, maximum=1.0):
    return f"{clamp_float(value, default, minimum, maximum):.2f}"


def build_preview_theme_data(request):
    data = {
        'key': 'preview',
        'name': (request.POST.get('theme_name') or 'Предпросмотр'),
        'font_family': request.POST.get('font_family') or 'Inter',
        'base_font_size': clamp_int(request.POST.get('base_font_size'), 16, 12, 28),
        'interface_scale': money_float(request.POST.get('interface_scale'), 1.25, 1.0, 2.2),
        'background_color': request.POST.get('theme_background_color') or '#eef6fb',
        'background_size': safe_size(request.POST.get('theme_background_size'), 'cover') if 'safe_size' in globals() else request.POST.get('theme_background_size') or 'cover',
        'background_position': safe_position(request.POST.get('theme_background_position'), 'center center') if 'safe_position' in globals() else request.POST.get('theme_background_position') or 'center center',
        'background_opacity': money_float(request.POST.get('theme_background_opacity'), 1.0, 0.0, 1.0),
        'overlay_effect': request.POST.get('theme_overlay_effect') or 'none',
        'overlay_opacity': money_float(request.POST.get('theme_overlay_opacity'), 0.35, 0.0, 1.0),
        'card_color': request.POST.get('theme_card_color') or '#ffffff',
        'card_transparency': money_float(request.POST.get('theme_card_transparency'), 0.88, 0.0, 1.0),
        'card_blur': clamp_int(request.POST.get('theme_card_blur'), 18, 0, 80),
        'card_style': request.POST.get('theme_card_style') or 'frosted',
        'card_border_opacity': money_float(request.POST.get('theme_card_border_opacity'), 0.24, 0.0, 1.0),
        'card_shadow_opacity': money_float(request.POST.get('theme_card_shadow_opacity'), 0.12, 0.0, 1.0),
        'reflection_strength': money_float(request.POST.get('theme_reflection_strength'), 0.25, 0.0, 1.0),
        'text_color': request.POST.get('theme_text_color') or '#17212b',
        'text_opacity': money_float(request.POST.get('theme_text_opacity'), 1.0, 0.15, 1.0),
        'muted_color': request.POST.get('theme_muted_color') or '#657786',
        'primary_color': request.POST.get('theme_primary_color') or '#229ED9',
        'danger_color': request.POST.get('theme_danger_color') or '#e0245e',
        'custom_css': request.POST.get('theme_custom_css') or '',
        'background_image': None,
        'background_gif': None,
        'overlay_gif': None,
    }
    return data


def ensure_default_theme_presets():
    try:
        for data in DEFAULT_THEME_PRESETS:
            obj, created = ThemePreset.objects.get_or_create(key=data['key'], defaults=data)
            if not created and obj.is_builtin:
                changed = False
                for field, value in data.items():
                    if hasattr(obj, field) and getattr(obj, field) != value:
                        setattr(obj, field, value)
                        changed = True
                if changed:
                    obj.save()
        ThemePreset.objects.filter(key__in=['anime-asuka', 'graphite-glass', 'silver-line', 'gif']).delete()
    except (OperationalError, ProgrammingError):
        # Таблица ещё не создана до миграции.
        pass


def get_theme_presets():
    ensure_default_theme_presets()
    try:
        return list(ThemePreset.objects.all().order_by('sort_order', 'name'))
    except (OperationalError, ProgrammingError):
        return []


def get_theme_by_key(key):
    ensure_default_theme_presets()
    try:
        return ThemePreset.objects.filter(key=key).first()
    except (OperationalError, ProgrammingError):
        return None


def normalize_theme(theme):
    if theme in ('gray-white', '', None, 'anime-asuka', 'gif', 'graphite-glass', 'silver-line'):
        theme = 'telegram-light'
    preset = get_theme_by_key(theme)
    if preset:
        return preset.key
    fallback = get_setting('default_theme', 'telegram-light')
    preset = get_theme_by_key(fallback)
    return preset.key if preset else 'telegram-light'


def active_theme_preset(theme_key):
    preset = get_theme_by_key(normalize_theme(theme_key))
    if preset:
        return preset
    # безопасный объект-заглушка до миграции
    return ThemePreset(key='telegram-light', name='Голубая белая')


def slugify_theme_name(name):
    base = slugify(name or 'theme', allow_unicode=False) or 'theme'
    base = base[:48]
    key = base
    i = 2
    while ThemePreset.objects.filter(key=key).exists():
        key = f'{base}-{i}'[:64]
        i += 1
    return key


def hex_to_rgb(hex_color):
    hex_color = (hex_color or '#ffffff').lstrip('#')
    if len(hex_color) != 6:
        return '255,255,255'
    try:
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        return f'{r},{g},{b}'
    except ValueError:
        return '255,255,255'


def safe_size(value, default='cover'):
    return value if value in ALLOWED_BACKGROUND_SIZES else default


def safe_position(value, default='center center'):
    return value if value in ALLOWED_BACKGROUND_POSITIONS else default


def apply_preset_settings(preset, request):
    preset.name = (request.POST.get('theme_name') or preset.name).strip() or preset.name
    preset.font_family = request.POST.get('font_family') or 'Inter'
    preset.base_font_size = clamp_int(request.POST.get('base_font_size'), 16, 12, 28)
    preset.interface_scale = money_float(request.POST.get('interface_scale'), 1.25, 1.0, 2.2)
    preset.background_color = request.POST.get('theme_background_color') or preset.background_color
    preset.background_size = safe_size(request.POST.get('theme_background_size'), 'cover')
    preset.background_position = safe_position(request.POST.get('theme_background_position'), 'center center')
    preset.background_opacity = money_float(request.POST.get('theme_background_opacity'), 1.0, 0.0, 1.0)
    preset.card_color = request.POST.get('theme_card_color') or preset.card_color
    preset.card_transparency = money_float(request.POST.get('theme_card_transparency'), 0.88, 0.0, 1.0)
    preset.card_blur = clamp_int(request.POST.get('theme_card_blur'), 18, 0, 80)
    preset.card_border_opacity = money_float(request.POST.get('theme_card_border_opacity'), 0.24, 0.0, 1.0)
    preset.card_shadow_opacity = money_float(request.POST.get('theme_card_shadow_opacity'), 0.12, 0.0, 1.0)
    preset.reflection_strength = money_float(request.POST.get('theme_reflection_strength'), 0.25, 0.0, 1.0)
    preset.text_color = request.POST.get('theme_text_color') or preset.text_color
    preset.text_opacity = money_float(request.POST.get('theme_text_opacity'), 1.0, 0.15, 1.0)
    preset.muted_color = request.POST.get('theme_muted_color') or preset.muted_color
    preset.primary_color = request.POST.get('theme_primary_color') or preset.primary_color
    preset.danger_color = request.POST.get('theme_danger_color') or preset.danger_color
    card_style = request.POST.get('theme_card_style') or 'frosted'
    if card_style in {key for key, _ in ThemePreset.CARD_STYLE_CHOICES}:
        preset.card_style = card_style
    preset.overlay_effect = request.POST.get('theme_overlay_effect') or 'none'
    preset.overlay_opacity = money_float(request.POST.get('theme_overlay_opacity'), 0.35, 0.0, 1.0)
    if 'theme_custom_css' in request.POST:
        preset.custom_css = request.POST.get('theme_custom_css') or ''
    if 'theme_overlay_gif' in request.FILES:
        preset.overlay_gif = request.FILES['theme_overlay_gif']
    if 'theme_background_image' in request.FILES:
        preset.background_image = request.FILES['theme_background_image']
    if 'theme_background_gif' in request.FILES:
        preset.background_gif = request.FILES['theme_background_gif']
    preset.save()
    return preset


def check_support_file(uploaded_file):
    if uploaded_file and uploaded_file.size > MAX_SUPPORT_FILE_SIZE:
        return 'Файл слишком большой. Максимальный размер — 5 МБ.'
    return ''


def touch_user_activity(request, force=False):
    if not request.user.is_authenticated:
        return None
    profile, _ = Profile.objects.get_or_create(user=request.user)
    now = timezone.now()
    if force or not profile.last_activity or profile.last_activity < now - timedelta(seconds=30):
        profile.last_activity = now
        profile.save(update_fields=['last_activity'])
    return profile


def admin_online_state():
    try:
        admin_profiles = Profile.objects.select_related('user').filter(user__is_active=True).filter(Q(user__username='admin') | Q(user__is_staff=True) | Q(user__is_superuser=True))
        threshold = timezone.now() - ONLINE_TIMEOUT
        return any(bool(p.last_activity and p.last_activity >= threshold) for p in admin_profiles)
    except Exception:
        return False


def get_user_support_thread(user):
    thread = SupportThread.objects.filter(user=user, kind='support', status='open').order_by('-updated_at').first()
    if not thread:
        thread = SupportThread.objects.create(user=user, kind='support', subject='Вопрос')
    return thread


def parse_date_cell(value):
    if not value:
        return None
    if hasattr(value, 'date'):
        return value.date()
    if isinstance(value, str):
        value = value.strip()
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y'):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return value


def normalize_choice(value, mapping, default):
    raw = (str(value or '').strip()).lower()
    if not raw:
        return default
    if raw in mapping:
        return mapping[raw]
    return default


def get_next_public_number():
    last = Olympiad.objects.exclude(public_number__isnull=True).order_by('-public_number').first()
    return (last.public_number + 1) if last else 0


def ensure_olympiad_public_numbers():
    next_number = get_next_public_number()
    for olympiad in Olympiad.objects.filter(public_number__isnull=True).order_by('id'):
        olympiad.public_number = next_number
        olympiad.save(update_fields=['public_number'])
        next_number += 1


def olympiad_matches_existing(public_number, title, start_date, end_date):
    if public_number is not None and Olympiad.objects.filter(public_number=public_number).exists():
        return True
    return Olympiad.objects.filter(title=title, start_date=start_date, end_date=end_date).exists()


def apply_generic_theme_settings(custom_theme, request):
    selected = normalize_theme(request.POST.get('selected_theme') or request.POST.get('edit_theme') or 'telegram-light')
    prefix_map = {
        'telegram-light': 'telegram',
        'custom': 'custom',
        'gif': 'gif',
        'anime-asuka': 'anime',
        'business': 'business',
    }
    prefix = prefix_map[selected]

    def set_attr(name, value):
        if hasattr(custom_theme, name):
            setattr(custom_theme, name, value)

    bg = request.POST.get('theme_background_color') or '#eef6fb'
    card = request.POST.get('theme_card_color') or '#ffffff'
    text = request.POST.get('theme_text_color') or '#17212b'
    muted = request.POST.get('theme_muted_color') or '#657786'
    primary = request.POST.get('theme_primary_color') or '#229ED9'
    card_transparency = f"{clamp_float(request.POST.get('theme_card_transparency'), 0.86, 0.0, 1.0):.2f}"
    text_opacity = f"{clamp_float(request.POST.get('theme_text_opacity'), 1.0, 0.15, 1.0):.2f}"
    overlay = f"{clamp_float(request.POST.get('theme_overlay_opacity'), 0.25, 0.0, 0.9):.2f}"
    bg_size = safe_size(request.POST.get('theme_background_size'), 'cover')
    bg_position = safe_position(request.POST.get('theme_background_position'), 'center center')
    card_blur = clamp_int(request.POST.get('theme_card_blur'), 18, 0, 60)
    card_style = request.POST.get('theme_card_style') or 'frosted'
    if card_style not in {'frosted', 'solid', 'outline', 'metallic', 'dark_glass'}:
        card_style = 'frosted'

    if selected == 'telegram-light':
        set_attr('telegram_background_color', bg)
        set_attr('telegram_card_color', card)
        set_attr('telegram_text_color', text)
        set_attr('telegram_muted_color', muted)
        set_attr('telegram_primary_color', primary)
        set_attr('telegram_card_transparency', card_transparency)
        set_attr('telegram_text_opacity', text_opacity)
        set_attr('telegram_overlay_opacity', overlay)
        set_attr('telegram_background_size', bg_size)
        set_attr('telegram_background_position', bg_position)
        set_attr('telegram_card_blur', card_blur)
        set_attr('telegram_card_style', card_style)
        if 'theme_background_image' in request.FILES:
            custom_theme.telegram_background_image = request.FILES['theme_background_image']
        if 'theme_background_gif' in request.FILES:
            custom_theme.telegram_background_gif = request.FILES['theme_background_gif']
    elif selected == 'custom':
        set_attr('background_color', bg)
        set_attr('card_color', card)
        set_attr('text_color', text)
        set_attr('muted_color', muted)
        set_attr('primary_color', primary)
        set_attr('btn_color', primary)
        set_attr('card_transparency', card_transparency)
        set_attr('text_opacity', text_opacity)
        set_attr('custom_overlay_opacity', overlay)
        set_attr('custom_background_size', bg_size)
        set_attr('custom_background_position', bg_position)
        set_attr('custom_card_blur', card_blur)
        set_attr('custom_card_style', card_style)
        custom_theme.custom_css = request.POST.get('custom_css') or ''
        if 'theme_background_image' in request.FILES:
            custom_theme.background_image = request.FILES['theme_background_image']
        if 'theme_background_gif' in request.FILES:
            custom_theme.custom_background_gif = request.FILES['theme_background_gif']
    elif selected == 'gif':
        set_attr('gif_text_color', text)
        set_attr('gif_primary_color', primary)
        set_attr('gif_card_transparency', card_transparency)
        set_attr('gif_text_opacity', text_opacity)
        set_attr('gif_overlay_opacity', overlay)
        set_attr('gif_size', bg_size)
        set_attr('gif_position', bg_position)
        set_attr('gif_card_blur', card_blur)
        set_attr('gif_card_style', card_style)
        if 'theme_background_gif' in request.FILES:
            custom_theme.background_gif = request.FILES['theme_background_gif']
    elif selected == 'anime-asuka':
        set_attr('anime_bg_color', bg)
        set_attr('anime_card_color', card)
        set_attr('anime_text_color', text)
        set_attr('anime_muted_color', muted)
        set_attr('anime_primary_color', primary)
        set_attr('anime_card_transparency', card_transparency)
        set_attr('anime_text_opacity', text_opacity)
        set_attr('anime_gif_overlay_opacity', overlay)
        set_attr('anime_gif_size', bg_size)
        set_attr('anime_gif_position', bg_position)
        set_attr('anime_card_blur', card_blur)
        set_attr('anime_card_style', card_style)
        custom_theme.anime_use_gif = 'theme_use_gif' in request.POST
        custom_theme.anime_hero_opacity = f"{clamp_float(request.POST.get('anime_hero_opacity'), 0.18, 0.0, 1.0):.2f}"
        if 'theme_background_image' in request.FILES:
            custom_theme.hero_image = request.FILES['theme_background_image']
        if 'theme_background_gif' in request.FILES:
            custom_theme.background_gif = request.FILES['theme_background_gif']
    elif selected == 'business':
        set_attr('business_background_color', bg)
        set_attr('business_card_color', card)
        set_attr('business_text_color', text)
        set_attr('business_muted_color', muted)
        set_attr('business_primary_color', primary)
        set_attr('business_card_transparency', card_transparency)
        set_attr('business_text_opacity', text_opacity)
        set_attr('business_overlay_opacity', overlay)
        set_attr('business_background_size', bg_size)
        set_attr('business_background_position', bg_position)
        set_attr('business_card_blur', card_blur)
        set_attr('business_card_style', card_style)
        custom_theme.business_reflection_strength = f"{clamp_float(request.POST.get('business_reflection_strength'), 0.34, 0.0, 1.0):.2f}"
        if 'theme_background_image' in request.FILES:
            custom_theme.business_background_image = request.FILES['theme_background_image']
        if 'theme_background_gif' in request.FILES:
            custom_theme.business_background_gif = request.FILES['theme_background_gif']
    return selected


def selected_theme_snapshot(custom_theme, selected):
    selected = normalize_theme(selected)
    if selected == 'custom':
        return {'background_color': custom_theme.background_color, 'card_color': custom_theme.card_color, 'text_color': custom_theme.text_color, 'muted_color': custom_theme.muted_color, 'primary_color': custom_theme.primary_color, 'card_transparency': custom_theme.card_transparency, 'text_opacity': custom_theme.text_opacity, 'overlay_opacity': custom_theme.custom_overlay_opacity, 'background_size': custom_theme.custom_background_size, 'background_position': custom_theme.custom_background_position, 'card_blur': custom_theme.custom_card_blur, 'card_style': custom_theme.custom_card_style}
    if selected == 'gif':
        return {'background_color': '#eef6fb', 'card_color': '#ffffff', 'text_color': custom_theme.gif_text_color, 'muted_color': '#657786', 'primary_color': custom_theme.gif_primary_color, 'card_transparency': custom_theme.gif_card_transparency, 'text_opacity': custom_theme.gif_text_opacity, 'overlay_opacity': custom_theme.gif_overlay_opacity, 'background_size': custom_theme.gif_size, 'background_position': custom_theme.gif_position, 'card_blur': custom_theme.gif_card_blur, 'card_style': custom_theme.gif_card_style}
    if selected == 'anime-asuka':
        return {'background_color': custom_theme.anime_bg_color, 'card_color': custom_theme.anime_card_color, 'text_color': custom_theme.anime_text_color, 'muted_color': custom_theme.anime_muted_color, 'primary_color': custom_theme.anime_primary_color, 'card_transparency': custom_theme.anime_card_transparency, 'text_opacity': custom_theme.anime_text_opacity, 'overlay_opacity': custom_theme.anime_gif_overlay_opacity, 'background_size': custom_theme.anime_gif_size, 'background_position': custom_theme.anime_gif_position, 'card_blur': custom_theme.anime_card_blur, 'card_style': custom_theme.anime_card_style}
    if selected == 'business':
        return {'background_color': custom_theme.business_background_color, 'card_color': custom_theme.business_card_color, 'text_color': custom_theme.business_text_color, 'muted_color': custom_theme.business_muted_color, 'primary_color': custom_theme.business_primary_color, 'card_transparency': custom_theme.business_card_transparency, 'text_opacity': custom_theme.business_text_opacity, 'overlay_opacity': custom_theme.business_overlay_opacity, 'background_size': custom_theme.business_background_size, 'background_position': custom_theme.business_background_position, 'card_blur': custom_theme.business_card_blur, 'card_style': custom_theme.business_card_style}
    return {'background_color': custom_theme.telegram_background_color, 'card_color': custom_theme.telegram_card_color, 'text_color': custom_theme.telegram_text_color, 'muted_color': custom_theme.telegram_muted_color, 'primary_color': custom_theme.telegram_primary_color, 'card_transparency': custom_theme.telegram_card_transparency, 'text_opacity': custom_theme.telegram_text_opacity, 'overlay_opacity': custom_theme.telegram_overlay_opacity, 'background_size': custom_theme.telegram_background_size, 'background_position': custom_theme.telegram_background_position, 'card_blur': custom_theme.telegram_card_blur, 'card_style': custom_theme.telegram_card_style}


def theme_context(request):
    custom_theme, _ = CustomTheme.objects.get_or_create(id=1)
    presets = get_theme_presets()
    default_theme = normalize_theme(get_setting('default_theme', 'telegram-light'))
    active_key = normalize_theme(request.session.get('theme', default_theme))
    preset = active_theme_preset(active_key)
    preview_mode = request.GET.get('preview_theme') == '1' and request.session.get('theme_preview_data')
    if preview_mode:
        preview = request.session.get('theme_preview_data') or {}
        preset = SimpleNamespace(**preview)

    support_thread = None
    support_messages = []
    user_support_unread = 0
    if request.user.is_authenticated:
        support_thread = SupportThread.objects.filter(user=request.user, status='open').order_by('-updated_at').first()
        if support_thread:
            support_messages = support_thread.messages.select_related('sender').order_by('created_at')[:80]
            user_support_unread = support_thread.messages.filter(is_admin=True, read_by_user=False).count()

    sound_enabled = request.session.get('sound_enabled')
    if sound_enabled is None:
        sound_enabled = '1' if custom_theme.sound_enabled_default else '0'

    return {
        'theme': preset.key,
        'default_theme': default_theme,
        'available_themes': presets,
        'theme_choices': [(p.key, p.name) for p in presets],
        'active_theme_preset': preset,
        'theme_preview_mode': preview_mode,
        'theme_preview_return_url': request.session.get('theme_preview_return_url', '/admin-panel/?tab=themes'),
        'theme_card_rgb': hex_to_rgb(preset.card_color),
        'theme_text_rgb': hex_to_rgb(preset.text_color),
        'theme_bg_rgb': hex_to_rgb(preset.background_color),
        'visually_impaired': request.session.get('visually_impaired', False),
        'sound_enabled': sound_enabled == '1',
        'custom_theme': custom_theme,
        'site_name': custom_theme.site_name or get_setting('site_name', 'Zauchka.RU'),
        'site_logo': custom_theme.site_logo.url if custom_theme.site_logo else '',
        'send_sound_url': custom_theme.send_sound.url if custom_theme.send_sound else '',
        'notification_sound_url': custom_theme.notification_sound.url if custom_theme.notification_sound else '',
        'report_sound_url': custom_theme.report_sound.url if hasattr(custom_theme, 'report_sound') and custom_theme.report_sound else '',
        'support_open_sound_url': custom_theme.support_open_sound.url if hasattr(custom_theme, 'support_open_sound') and custom_theme.support_open_sound else '',
        'support_close_sound_url': custom_theme.support_close_sound.url if hasattr(custom_theme, 'support_close_sound') and custom_theme.support_close_sound else '',
        'send_sound_volume': getattr(custom_theme, 'send_sound_volume', 100),
        'notification_sound_volume': getattr(custom_theme, 'notification_sound_volume', 100),
        'report_sound_volume': getattr(custom_theme, 'report_sound_volume', 100),
        'support_open_sound_volume': getattr(custom_theme, 'support_open_sound_volume', 100),
        'support_close_sound_volume': getattr(custom_theme, 'support_close_sound_volume', 100),
        'support_widget_gif': custom_theme.support_widget_gif.url if custom_theme.support_widget_gif else None,
        'support_widget_text': custom_theme.support_widget_text or 'Привет, буду рад помочь!',
        'support_widget_title': custom_theme.support_widget_title or 'Помощь администратора',
        'support_thread': support_thread,
        'support_messages': support_messages,
        'user_support_unread': user_support_unread,
        'admin_is_online': admin_online_state(),
        'show_cookie_banner': not request.user.is_authenticated and not request.COOKIES.get('cookies_accepted'),
        'footer_text': get_setting('footer_text', 'Разработчик: m143eo@gmail.com'),
        'rules_content': get_setting('rules_content', DEFAULT_RULES_CONTENT),
        'cookies_content': get_setting('cookies_content', DEFAULT_COOKIES_CONTENT),
        'russian_regions': RUSSIAN_REGIONS,
    }


def render_main(request, template_name, context=None):
    touch_user_activity(request)
    data = theme_context(request)
    if context:
        data.update(context)
    return render(request, template_name, data)


def index(request):
    ensure_olympiad_public_numbers()
    olympiads = Olympiad.objects.filter(is_approved=True).order_by('-date_added')
    target = request.GET.get('target')
    grade = request.GET.get('grade')
    stage = request.GET.get('stage')
    format_type = request.GET.get('format')
    region = request.GET.get('region')
    listing_type = request.GET.get('listing_type')
    q = (request.GET.get('q') or '').strip()

    if target:
        olympiads = olympiads.filter(target_audience=target)
    if grade:
        olympiads = olympiads.filter(grade=grade)
    if stage:
        olympiads = olympiads.filter(stage=stage)
    if format_type:
        olympiads = olympiads.filter(format=format_type)
    if region:
        olympiads = olympiads.filter(region__icontains=region)
    if listing_type:
        olympiads = olympiads.filter(listing_type=listing_type)
    if q:
        clean_code = q.replace('#', '').strip()
        query = Q(title__icontains=q) | Q(short_description__icontains=q) | Q(description__icontains=q) | Q(region__icontains=q) | Q(organizer__icontains=q) | Q(stage__icontains=q) | Q(format__icontains=q) | Q(listing_type__icontains=q)
        if clean_code.isdigit():
            query |= Q(public_number=int(clean_code))
        if 'грант' in q.lower():
            query |= Q(listing_type='grant')
        if 'олимпиад' in q.lower():
            query |= Q(listing_type='olympiad')
        olympiads = olympiads.filter(query).distinct()

    active_olympiads = [o for o in olympiads if o.is_active]
    archive_olympiads = [o for o in olympiads if not o.is_active]

    sort_by = request.GET.get('sort')

    favorite_ids = set()
    favorite_order = {}
    if request.user.is_authenticated:
        favorites = list(FavoriteItem.objects.filter(user=request.user).select_related('olympiad').order_by('-created_at'))
        favorite_ids = {f.olympiad_id for f in favorites}
        favorite_order = {f.olympiad_id: i for i, f in enumerate(favorites)}
    for item in list(active_olympiads) + list(archive_olympiads):
        item.is_favorite = item.id in favorite_ids

    if sort_by == 'name':
        active_olympiads.sort(key=lambda x: x.title)
        archive_olympiads.sort(key=lambda x: x.title)
    elif sort_by == 'deadline':
        active_olympiads.sort(key=lambda x: x.end_date)
        archive_olympiads.sort(key=lambda x: x.end_date)
    elif sort_by == 'favorites' and request.user.is_authenticated:
        active_olympiads.sort(key=lambda x: (0 if x.id in favorite_ids else 1, favorite_order.get(x.id, 10**9), x.title.lower()))
        archive_olympiads.sort(key=lambda x: (0 if x.id in favorite_ids else 1, favorite_order.get(x.id, 10**9), x.title.lower()))

    return render_main(request, 'main/index.html', {
        'active': active_olympiads,
        'archive': archive_olympiads,
        'search_query': q,
        'sort_by': sort_by,
        'favorite_ids': favorite_ids,
        'is_real_admin': is_site_admin(request.user),
        'show_cookie_banner': not request.user.is_authenticated and not request.COOKIES.get('cookies_accepted'),
    })


def olympiad_detail(request, pk):
    olympiad = get_object_or_404(Olympiad, pk=pk, is_approved=True)
    comments = list(olympiad.comments.filter(is_approved=True).select_related('user', 'deleted_by').order_by('-created_at'))
    for comment in comments:
        Profile.objects.get_or_create(user=comment.user)
    is_favorite = request.user.is_authenticated and FavoriteItem.objects.filter(user=request.user, olympiad=olympiad).exists()
    return render_main(request, 'main/olympiad_detail.html', {
        'olympiad': olympiad,
        'comments': comments,
        'is_favorite': is_favorite,
        'is_real_admin': is_site_admin(request.user),
    })


def agreement_view(request):
    next_url = request.META.get('HTTP_REFERER', '/')
    return render_main(request, 'main/agreement.html', {'next_url': next_url})


def cookies_view(request):
    next_url = request.META.get('HTTP_REFERER', '/')
    return render_main(request, 'main/cookies.html', {'next_url': next_url})


def add_comment(request, pk):
    if not request.user.is_authenticated:
        return redirect('login')

    olympiad = get_object_or_404(Olympiad, pk=pk)
    profile, _ = Profile.objects.get_or_create(user=request.user)
    if not profile.can_comment:
        return redirect('olympiad_detail', pk=pk)
    text = (request.POST.get('text') or '').strip()
    if not text:
        return redirect('olympiad_detail', pk=pk)

    is_approved = True
    for bw in BannedWord.objects.all():
        if bw.word.lower() in text.lower():
            is_approved = False
            break

    Comment.objects.create(olympiad=olympiad, user=request.user, text=text, is_approved=is_approved)
    return redirect('olympiad_detail', pk=pk)


def report_comment(request, pk):
    if not request.user.is_authenticated:
        return redirect('login')

    comment = get_object_or_404(Comment.objects.select_related('olympiad', 'user'), pk=pk)
    text = (request.POST.get('report_text') or '').strip() or 'Пользователь отправил жалобу на комментарий.'

    thread = SupportThread.objects.create(
        user=request.user,
        kind='comment_report',
        comment=comment,
        subject=f'Жалоба на комментарий пользователя {comment.user.username}'
    )
    SupportMessage.objects.create(
        thread=thread,
        sender=request.user,
        text=f'{text}\n\nЖалоба на комментарий пользователя {comment.user.username}: {comment.text}',
        is_admin=False,
        read_by_user=True,
    )
    return redirect('olympiad_detail', pk=comment.olympiad.pk)



@require_POST
def toggle_favorite(request, pk):
    if not request.user.is_authenticated:
        return redirect('login')
    olympiad = get_object_or_404(Olympiad, pk=pk, is_approved=True)
    existing = FavoriteItem.objects.filter(user=request.user, olympiad=olympiad).first()
    if existing:
        existing.delete()
    else:
        FavoriteItem.objects.create(user=request.user, olympiad=olympiad)
    return redirect(request.META.get('HTTP_REFERER', 'index'))

def delete_comment(request, pk):
    if request.user.is_authenticated:
        Comment.objects.filter(pk=pk, user=request.user).delete()
        return redirect('profile', pk=request.user.pk)
    return redirect('login')


def register_view(request):
    if request.method == 'POST':
        username = (request.POST.get('username') or '').strip()
        password = request.POST.get('password')
        email = (request.POST.get('email') or '').strip()
        pd_agree = request.POST.get('pd_agree') == 'on'
        cookies_agree = request.POST.get('cookies_agree') == 'on'

        if not pd_agree or not cookies_agree:
            return render_main(request, 'main/register.html', {'error': 'Необходимо принять соглашение и cookies'})
        if User.objects.filter(username=username).exists():
            return render_main(request, 'main/register.html', {'error': 'Такой логин уже занят'})

        user = User.objects.create_user(username=username, password=password, email=email)
        profile = Profile.objects.create(user=user, pd_agreement=True, cookies_agreement=True)
        if profile.can_upload_photo and 'photo' in request.FILES:
            profile.photo = request.FILES['photo']
        profile.region = request.POST.get('region') or ''
        profile.save()

        login(request, user)
        touch_user_activity(request, force=True)
        return redirect('index')

    return render_main(request, 'main/register.html')


def login_view(request):
    if request.method == 'POST':
        user = authenticate(username=request.POST.get('username'), password=request.POST.get('password'))
        if user:
            profile, _ = Profile.objects.get_or_create(user=user)
            if profile.is_banned:
                if profile.ban_end_date and profile.ban_end_date < timezone.now():
                    profile.is_banned = False
                    profile.save()
                else:
                    end_str = f" до {profile.ban_end_date.strftime('%d.%m.%Y %H:%M')}" if profile.ban_end_date else " навсегда"
                    return render_main(request, 'main/login.html', {'ban_error': f'Аккаунт заблокирован{end_str}. Причина: {profile.ban_reason}'})

            profile.ip_address = request.META.get('REMOTE_ADDR')
            profile.last_login = timezone.now()
            profile.last_activity = timezone.now()
            profile.save()
            login(request, user)
            return redirect('index')
        return render_main(request, 'main/login.html', {'error': 'Неверный логин или пароль'})

    return render_main(request, 'main/login.html')


def logout_view(request):
    if request.user.is_authenticated:
        Profile.objects.filter(user=request.user).update(last_activity=None)
    logout(request)
    return redirect('index')


def profile_view(request, pk):
    user = get_object_or_404(User, pk=pk)
    profile, created = Profile.objects.get_or_create(user=user)
    is_owner = (request.user == user)
    is_admin_viewing = is_site_admin(request.user)
    user_comments = Comment.objects.filter(user=user).select_related('olympiad', 'deleted_by').order_by('-created_at') if is_owner or is_admin_viewing else None

    return render_main(request, 'main/profile.html', {
        'u': user,
        'profile': profile,
        'is_owner': is_owner,
        'comments': user_comments,
    })


def edit_profile_view(request):
    if not request.user.is_authenticated:
        return redirect('login')

    profile, _ = Profile.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        if profile.can_change_username:
            new_username = (request.POST.get('username') or request.user.username).strip()
            if new_username and new_username != request.user.username and not User.objects.filter(username=new_username).exclude(pk=request.user.pk).exists():
                request.user.username = new_username
                request.user.save(update_fields=['username'])
        profile.bio = request.POST.get('bio') or ''
        profile.region = request.POST.get('region') or ''
        profile.show_region = 'show_region' in request.POST
        profile.show_last_login = 'show_last_login' in request.POST
        profile.show_online_status = 'show_online_status' in request.POST
        if profile.can_upload_photo and 'photo' in request.FILES:
            profile.photo = request.FILES['photo']
        profile.save()
        return redirect('profile', pk=request.user.pk)

    return render_main(request, 'main/edit_profile.html')


def support_center(request):
    if not request.user.is_authenticated:
        return redirect('login')

    thread = get_user_support_thread(request.user)
    error = ''
    if request.method == 'POST':
        text = (request.POST.get('text') or '').strip()
        uploaded = request.FILES.get('file')
        error = check_support_file(uploaded)
        if not error and (text or uploaded):
            SupportMessage.objects.create(thread=thread, sender=request.user, text=text, file=uploaded, is_admin=False, read_by_user=True)
            thread.status = 'open'
            thread.save(update_fields=['status', 'updated_at'])
            return redirect(request.META.get('HTTP_REFERER', 'index'))

    thread.messages.filter(is_admin=True, read_by_user=False).update(read_by_user=True)
    return render_main(request, 'main/support.html', {
        'support_thread': thread,
        'support_messages': thread.messages.select_related('sender'),
        'support_error': error,
    })


@require_POST
def online_ping(request):
    if request.user.is_authenticated:
        touch_user_activity(request, force=True)
        return JsonResponse({'ok': True, 'online': True})
    return JsonResponse({'ok': False, 'online': False}, status=401)


def login_admin(request):
    if request.method == 'POST':
        user = authenticate(username=request.POST.get('username'), password=request.POST.get('password'))
        if user and (user.username == 'admin' or user.is_staff or user.is_superuser):
            login(request, user)
            request.session['is_admin'] = True
            touch_user_activity(request, force=True)
            return redirect('admin_panel')
        return render_main(request, 'main/login_admin.html', {'error': 'Неверные данные администратора'})
    return render_main(request, 'main/login_admin.html')



def build_media_inventory():
    """Return all files from MEDIA_ROOT with best-effort usage detection.

    The previous version only checked a few fields manually. This version also
    scans every FileField/ImageField in the app, so GIF wallpapers, overlay GIFs,
    chat attachments, profile photos, olympiad covers and future media fields are
    all visible in "Архив данных".
    """
    media_root = getattr(settings, 'MEDIA_ROOT', '')
    rows = []
    if not media_root or not os.path.isdir(media_root):
        return rows

    usage = {}

    def normalize_name(name):
        return (name or '').replace('\\', '/').lstrip('/')

    def add_usage(field_file, label):
        try:
            name = normalize_name(getattr(field_file, 'name', '') or '')
            if name:
                usage.setdefault(name, set()).add(label)
        except Exception:
            pass

    def describe_instance(obj, field_name):
        model = obj.__class__.__name__
        if isinstance(obj, Profile):
            return f'Фото профиля: {obj.user.username}'
        if isinstance(obj, Olympiad):
            return f'Олимпиада: {obj.title}'
        if isinstance(obj, ThemePreset):
            return f'Тема: {obj.name} · {field_name}'
        if isinstance(obj, CustomTheme):
            return f'Общие настройки сайта · {field_name}'
        if isinstance(obj, SupportMessage):
            sender = obj.sender.username if obj.sender else 'удалённый пользователь'
            return f'Чат #{obj.thread_id}: файл сообщения от {sender}'
        return f'{model} · {field_name}'

    # Generic scan for all FileField/ImageField values in the main app.
    try:
        from django.apps import apps
        from django.db.models import FileField
        for model in apps.get_app_config('main').get_models():
            file_fields = [f for f in model._meta.get_fields() if isinstance(f, FileField)]
            if not file_fields:
                continue
            qs = model.objects.all()
            for obj in qs.iterator():
                for field in file_fields:
                    add_usage(getattr(obj, field.name, None), describe_instance(obj, field.verbose_name or field.name))
    except Exception:
        # Fallback: preserve the old critical checks if generic introspection fails.
        for p in Profile.objects.select_related('user'):
            add_usage(p.photo, f'Фото профиля: {p.user.username}')
        for o in Olympiad.objects.all():
            add_usage(o.cover_image, f'Обложка: {o.title}')
        for preset in ThemePreset.objects.all():
            add_usage(preset.background_image, f'Фон темы: {preset.name}')
            add_usage(preset.background_gif, f'GIF темы: {preset.name}')
            add_usage(preset.overlay_gif, f'GIF поверх сайта: {preset.name}')
        custom, _ = CustomTheme.objects.get_or_create(id=1)
        for field, label in [
            ('site_logo','Логотип сайта'),('send_sound','Звук отправки'),
            ('notification_sound','Звук уведомления'),('report_sound','Звук жалобы'),
            ('support_open_sound','Звук открытия помощи'),('support_close_sound','Звук закрытия помощи'),
            ('support_widget_gif','GIF окна помощи'),('background_gif','GIF фон'),
            ('custom_background_gif','GIF оригинальной темы'),('business_background_gif','GIF деловой темы')
        ]:
            add_usage(getattr(custom, field, None), label)
        for m in SupportMessage.objects.select_related('sender'):
            sender = m.sender.username if m.sender else 'удалённый пользователь'
            add_usage(m.file, f'Чат #{m.thread_id}: файл сообщения от {sender}')

    for root, _, files in os.walk(media_root):
        for file_name in files:
            full = os.path.join(root, file_name)
            rel = os.path.relpath(full, media_root).replace('\\','/')
            mime, _ = mimetypes.guess_type(full)
            rows.append({
                'name': file_name,
                'path': rel,
                'size': os.path.getsize(full),
                'mime': mime or 'file',
                'usage': sorted(usage.get(rel, [])),
            })
    rows.sort(key=lambda x: (0 if x['usage'] else 1, -x['size']))
    return rows

def admin_panel(request):
    is_admin = is_site_admin(request.user)
    if not is_admin and not request.session.get('is_admin'):
        return redirect('login_admin')

    touch_user_activity(request, force=True)
    tab = request.GET.get('tab', 'active')
    olymp_to_edit = None
    custom_theme, _ = CustomTheme.objects.get_or_create(id=1)

    if tab == 'edit':
        edit_id = request.GET.get('id')
        if edit_id:
            olymp_to_edit = get_object_or_404(Olympiad, id=edit_id)

    selected_thread = None
    selected_thread_id = request.GET.get('thread')
    if selected_thread_id:
        selected_thread = get_object_or_404(SupportThread.objects.select_related('user', 'comment'), id=selected_thread_id)
        selected_thread.messages.filter(is_admin=False, read_by_admin=False).update(read_by_admin=True)

    if request.method == 'POST':
        if 'approve_olymp' in request.POST:
            Olympiad.objects.filter(id=request.POST.get('approve_olymp')).update(is_approved=True)
            return redirect('/admin-panel/?tab=pending')

        elif 'reject_olymp' in request.POST:
            o = get_object_or_404(Olympiad, id=request.POST.get('reject_olymp'))
            o.is_approved = True
            o.end_date = timezone.localdate() - timedelta(days=1)
            o.save(update_fields=['is_approved', 'end_date'])
            return redirect('/admin-panel/?tab=pending')

        elif 'delete_olymp' in request.POST:
            Olympiad.objects.filter(id=request.POST.get('delete_olymp')).delete()
            return redirect('/admin-panel/?tab=active')

        elif 'approve_comment' in request.POST:
            Comment.objects.filter(id=request.POST.get('approve_comment')).update(is_approved=True)
            return redirect('/admin-panel/?tab=comments')

        elif 'delete_comment_site' in request.POST:
            comment = get_object_or_404(Comment, id=request.POST.get('delete_comment_site'))
            comment.is_deleted = True
            comment.deleted_by = request.user
            comment.deleted_at = timezone.now()
            comment.delete_reason = request.POST.get('delete_reason') or ''
            comment.show_delete_notice = 'hide_delete_notice' not in request.POST
            comment.save(update_fields=['is_deleted', 'deleted_by', 'deleted_at', 'delete_reason', 'show_delete_notice'])
            return redirect('/admin-panel/?tab=comments')

        elif 'delete_comment' in request.POST or 'admin_del_comment' in request.POST:
            comment_id = request.POST.get('delete_comment') or request.POST.get('admin_del_comment')
            comment = get_object_or_404(Comment, id=comment_id)
            if is_site_admin(request.user) and comment.user != request.user:
                comment.is_deleted = True
                comment.deleted_by = request.user
                comment.deleted_at = timezone.now()
                comment.delete_reason = request.POST.get('delete_reason') or ''
                comment.show_delete_notice = 'hide_delete_notice' not in request.POST
                comment.save(update_fields=['is_deleted', 'deleted_by', 'deleted_at', 'delete_reason', 'show_delete_notice'])
            elif comment.user == request.user:
                comment.delete()
            return redirect(request.META.get('HTTP_REFERER', '/admin-panel/?tab=active'))

        elif 'hide_comment_delete_notice' in request.POST:
            comment = get_object_or_404(Comment, id=request.POST.get('hide_comment_delete_notice'))
            comment.show_delete_notice = False
            comment.save(update_fields=['show_delete_notice'])
            return redirect(request.META.get('HTTP_REFERER', '/admin-panel/?tab=active'))

        elif 'add_word' in request.POST:
            word = (request.POST.get('word') or '').strip()
            if word:
                BannedWord.objects.get_or_create(word=word)
            return redirect('/admin-panel/?tab=words')

        elif 'delete_word' in request.POST:
            BannedWord.objects.filter(id=request.POST.get('delete_word')).delete()
            return redirect('/admin-panel/?tab=words')

        elif 'ban_user' in request.POST:
            p = Profile.objects.get(user_id=request.POST.get('ban_user'))
            p.is_banned = True
            p.ban_reason = request.POST.get('ban_reason') or 'Причина не указана'
            p.ban_end_date = request.POST.get('ban_end_date') or None
            if p.ban_end_date:
                p.ban_end_date = datetime.strptime(p.ban_end_date, '%Y-%m-%dT%H:%M')
            p.save()
            return redirect('/admin-panel/?tab=users')

        elif 'unban_user' in request.POST:
            p = Profile.objects.get(user_id=request.POST.get('unban_user'))
            p.is_banned = False
            p.ban_reason = ''
            p.ban_end_date = None
            p.save()
            return redirect('/admin-panel/?tab=users')

        elif 'delete_user' in request.POST:
            user_id = request.POST.get('delete_user')
            target = get_object_or_404(User, id=user_id)
            if target.username != 'admin' and target != request.user:
                target.delete()
            return redirect('/admin-panel/?tab=users')

        elif 'set_admin_rights' in request.POST or 'update_user_permissions' in request.POST:
            target_id = request.POST.get('set_admin_rights') or request.POST.get('update_user_permissions')
            target = get_object_or_404(User, id=target_id)
            profile, _ = Profile.objects.get_or_create(user=target)
            if target.username != 'admin':
                make_admin = request.POST.get('make_admin') == 'on'
                target.is_staff = make_admin
                target.is_superuser = make_admin
                target.save(update_fields=['is_staff', 'is_superuser'])
            profile.can_upload_photo = 'can_upload_photo' in request.POST
            profile.can_change_username = 'can_change_username' in request.POST
            profile.can_comment = 'can_comment' in request.POST
            profile.save(update_fields=['can_upload_photo', 'can_change_username', 'can_comment'])
            return redirect('/admin-panel/?tab=users')

        elif 'change_user_password' in request.POST:
            target = get_object_or_404(User, id=request.POST.get('change_user_password'))
            changed_fields = []
            new_username = (request.POST.get('new_username') or '').strip()
            if new_username and new_username != target.username and not User.objects.filter(username=new_username).exclude(pk=target.pk).exists():
                target.username = new_username
                changed_fields.append('username')
            new_password = request.POST.get('new_password') or ''
            if len(new_password) >= 4:
                target.password = make_password(new_password)
                changed_fields.append('password')
            if changed_fields:
                target.save(update_fields=changed_fields)
            return redirect('/admin-panel/?tab=users')

        elif 'admin_message_user' in request.POST:
            target = get_object_or_404(User, id=request.POST.get('admin_message_user'))
            text = (request.POST.get('admin_message_text') or '').strip()
            thread = SupportThread.objects.filter(user=target, kind='support', status='open').order_by('-updated_at').first()
            if not thread:
                thread = SupportThread.objects.create(user=target, kind='support', subject='Вопрос')
            if text:
                SupportMessage.objects.create(thread=thread, sender=request.user, text=text, is_admin=True, read_by_admin=True)
                thread.save(update_fields=['updated_at'])
            return redirect(f'/admin-panel/?tab=appeals&thread={thread.id}')

        elif 'delete_support_file' in request.POST:
            msg = get_object_or_404(SupportMessage, id=request.POST.get('delete_support_file'))
            thread_id = msg.thread_id
            if msg.file:
                msg.file.delete(save=False)
                msg.file = ''
                msg.save(update_fields=['file'])
            return redirect(f'/admin-panel/?tab=appeals&thread={thread_id}')

        elif 'delete_media_file' in request.POST:
            rel = (request.POST.get('delete_media_file') or '').replace('\\', '/').lstrip('/')
            abs_path = os.path.abspath(os.path.join(settings.MEDIA_ROOT, rel))
            media_root = os.path.abspath(settings.MEDIA_ROOT)
            if abs_path.startswith(media_root) and os.path.exists(abs_path):
                os.remove(abs_path)
            return redirect('/admin-panel/?tab=media')

        elif 'create_theme' in request.POST:
            name = (request.POST.get('new_theme_name') or 'Новая тема').strip()
            key = slugify_theme_name(name)
            source = active_theme_preset(request.POST.get('copy_from_theme') or get_setting('default_theme', 'telegram-light'))
            preset = ThemePreset.objects.create(
                key=key,
                name=name,
                is_builtin=False,
                sort_order=(ThemePreset.objects.count() + 1) * 10,
                font_family=source.font_family,
                base_font_size=source.base_font_size,
                interface_scale=source.interface_scale,
                background_color=source.background_color,
                background_size=source.background_size,
                background_position=source.background_position,
                background_opacity=source.background_opacity,
                card_color=source.card_color,
                card_transparency=source.card_transparency,
                card_blur=source.card_blur,
                card_style=source.card_style,
                card_border_opacity=source.card_border_opacity,
                card_shadow_opacity=source.card_shadow_opacity,
                reflection_strength=source.reflection_strength,
                text_color=source.text_color,
                text_opacity=source.text_opacity,
                muted_color=source.muted_color,
                primary_color=source.primary_color,
                danger_color=source.danger_color,
                custom_css=source.custom_css,
            )
            return redirect(f'/admin-panel/?tab=themes&edit_theme={preset.key}')

        elif 'delete_theme' in request.POST:
            key = request.POST.get('delete_theme')
            preset = get_object_or_404(ThemePreset, key=key)
            if not preset.is_builtin and ThemePreset.objects.count() > 1:
                preset.delete()
                if get_setting('default_theme', 'telegram-light') == key:
                    set_setting('default_theme', 'telegram-light')
            return redirect('/admin-panel/?tab=themes')

        elif 'preview_theme_editor' in request.POST:
            selected_key = normalize_theme(request.POST.get('selected_theme') or 'telegram-light')
            request.session['theme_preview_data'] = build_preview_theme_data(request)
            request.session['theme_preview_return_url'] = f'/admin-panel/?tab=themes&edit_theme={selected_key}'
            return redirect('/?preview_theme=1')

        elif 'cancel_theme_preview' in request.POST:
            request.session.pop('theme_preview_data', None)
            request.session.pop('theme_preview_return_url', None)
            selected_key = normalize_theme(request.POST.get('selected_theme') or 'telegram-light')
            return redirect(f'/admin-panel/?tab=themes&edit_theme={selected_key}')

        elif 'save_theme_editor' in request.POST:
            selected_key = normalize_theme(request.POST.get('selected_theme') or 'telegram-light')
            preset = get_object_or_404(ThemePreset, key=selected_key)
            apply_preset_settings(preset, request)
            request.session.pop('theme_preview_data', None)
            request.session.pop('theme_preview_return_url', None)
            return redirect(f'/admin-panel/?tab=themes&edit_theme={preset.key}')

        elif 'save_site_common' in request.POST:
            common_selected = request.POST.get('selected_theme') or get_setting('default_theme', 'telegram-light')
            custom_theme.site_name = request.POST.get('site_name') or 'Zauchka.RU'
            custom_theme.sound_enabled_default = 'sound_enabled_default' in request.POST
            custom_theme.support_widget_text = request.POST.get('support_widget_text') or 'Привет, буду рад помочь!'
            custom_theme.support_widget_title = request.POST.get('support_widget_title') or 'Помощь администратора'
            if 'site_logo' in request.FILES:
                custom_theme.site_logo = request.FILES['site_logo']
            if 'support_widget_gif' in request.FILES:
                custom_theme.support_widget_gif = request.FILES['support_widget_gif']
            if 'send_sound' in request.FILES:
                custom_theme.send_sound = request.FILES['send_sound']
            if 'notification_sound' in request.FILES:
                custom_theme.notification_sound = request.FILES['notification_sound']
            if 'report_sound' in request.FILES:
                custom_theme.report_sound = request.FILES['report_sound']
            if 'support_open_sound' in request.FILES:
                custom_theme.support_open_sound = request.FILES['support_open_sound']
            if 'support_close_sound' in request.FILES:
                custom_theme.support_close_sound = request.FILES['support_close_sound']
            custom_theme.send_sound_volume = clamp_int(request.POST.get('send_sound_volume'), 100, 0, 200)
            custom_theme.notification_sound_volume = clamp_int(request.POST.get('notification_sound_volume'), 100, 0, 200)
            custom_theme.report_sound_volume = clamp_int(request.POST.get('report_sound_volume'), 100, 0, 200)
            custom_theme.support_open_sound_volume = clamp_int(request.POST.get('support_open_sound_volume'), 100, 0, 200)
            custom_theme.support_close_sound_volume = clamp_int(request.POST.get('support_close_sound_volume'), 100, 0, 200)
            custom_theme.save()
            set_setting('default_theme', normalize_theme(request.POST.get('default_theme', get_setting('default_theme', 'telegram-light'))))
            set_setting('site_name', custom_theme.site_name)
            set_setting('footer_text', request.POST.get('footer_text', ''))
            return redirect(f'/admin-panel/?tab=themes&edit_theme={common_selected}')

        elif 'update_thread_title' in request.POST:
            thread = get_object_or_404(SupportThread, id=request.POST.get('thread_id'))
            thread.subject = (request.POST.get('subject') or thread.get_kind_display()).strip()
            thread.save(update_fields=['subject', 'updated_at'])
            return redirect(f'/admin-panel/?tab=appeals&thread={thread.id}')

        elif 'admin_support_reply' in request.POST:
            thread = get_object_or_404(SupportThread, id=request.POST.get('thread_id'))
            text = (request.POST.get('text') or '').strip()
            uploaded = request.FILES.get('file')
            if not check_support_file(uploaded) and (text or uploaded):
                SupportMessage.objects.create(thread=thread, sender=request.user, text=text, file=uploaded, is_admin=True, read_by_admin=True)
                thread.status = 'open'
                thread.save(update_fields=['status', 'updated_at'])
            return redirect(f'/admin-panel/?tab=appeals&thread={thread.id}')

        elif 'close_thread' in request.POST:
            thread = get_object_or_404(SupportThread, id=request.POST.get('close_thread'))
            thread.status = 'closed'
            thread.save(update_fields=['status', 'updated_at'])
            return redirect(f'/admin-panel/?tab=appeals&thread={thread.id}')

        elif 'open_thread' in request.POST:
            thread = get_object_or_404(SupportThread, id=request.POST.get('open_thread'))
            thread.status = 'open'
            thread.save(update_fields=['status', 'updated_at'])
            return redirect(f'/admin-panel/?tab=appeals&thread={thread.id}')

        elif 'update_olympiad_cover' in request.POST:
            o = get_object_or_404(Olympiad, id=request.POST.get('olympiad_id'))
            if 'cover_image' in request.FILES:
                o.cover_image = request.FILES['cover_image']
                o.save(update_fields=['cover_image'])
            return redirect(request.META.get('HTTP_REFERER', '/admin-panel/?tab=active'))

        elif 'edit_id' in request.POST:
            o = Olympiad.objects.get(id=request.POST.get('edit_id'))
            o.title = request.POST.get('title')
            o.listing_type = request.POST.get('listing_type') or 'olympiad'
            o.organizer = request.POST.get('organizer') or ''
            o.link = request.POST.get('link')
            o.start_date = request.POST.get('start_date')
            o.end_date = request.POST.get('end_date')
            o.target_audience = request.POST.get('target_audience')
            o.grade = request.POST.get('grade') or None
            o.stage = request.POST.get('stage')
            o.format = request.POST.get('format')
            o.region = request.POST.get('region')
            o.short_description = request.POST.get('short_description') or ''
            o.description = request.POST.get('description')
            if 'cover_image' in request.FILES:
                o.cover_image = request.FILES['cover_image']
            o.save()
            return redirect('/admin-panel/?tab=edit&id=' + str(o.id))

        elif 'excel_file' in request.FILES:
            file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [str(cell.value or '').strip().lower() for cell in ws[1]]
            def cell(row, names, fallback_index=None):
                for name in names:
                    if name.lower() in headers:
                        return row[headers.index(name.lower())]
                if fallback_index is not None and fallback_index < len(row):
                    return row[fallback_index]
                return None
            target_map = {'school': 'school', 'школьник': 'school', 'student': 'student', 'студент': 'student'}
            stage_map = {'municipal': 'municipal', 'муниципальный': 'municipal', 'regional': 'regional', 'региональный': 'regional', 'all_russia': 'all_russia', 'всероссийский': 'all_russia'}
            format_map = {'online': 'online', 'онлайн': 'online', 'заочно (онлайн)': 'online', 'offline': 'offline', 'очно': 'offline'}
            type_map = {'olympiad': 'olympiad', 'олимпиада': 'olympiad', 'grant': 'grant', 'грант': 'grant'}
            for row in ws.iter_rows(min_row=2, values_only=True):
                title = cell(row, ['Название', 'title'], 0)
                if not title:
                    continue
                raw_number = cell(row, ['Номер', 'ID', 'Код', '#'])
                public_number = None
                if raw_number not in (None, ''):
                    raw = str(raw_number).replace('#', '').strip()
                    if raw.isdigit():
                        public_number = int(raw)
                start_date = parse_date_cell(cell(row, ['Начало', 'start_date'], 3))
                end_date = parse_date_cell(cell(row, ['Конец', 'end_date'], 4))
                if not start_date or not end_date or olympiad_matches_existing(public_number, title, start_date, end_date):
                    continue
                Olympiad.objects.create(
                    public_number=public_number,
                    listing_type=normalize_choice(cell(row, ['Тип', 'type'], 1), type_map, 'olympiad'),
                    title=title,
                    organizer=cell(row, ['Организатор', 'От кого', 'organizer'], 2) or '',
                    link=cell(row, ['Ссылка', 'link'], 5) or 'https://example.com',
                    start_date=start_date,
                    end_date=end_date,
                    target_audience=normalize_choice(cell(row, ['Аудитория', 'target_audience'], 6), target_map, 'school'),
                    grade=cell(row, ['Класс', 'grade'], 7) or None,
                    stage=normalize_choice(cell(row, ['Уровень', 'Этап', 'stage'], 8), stage_map, 'municipal'),
                    format=normalize_choice(cell(row, ['Формат', 'format'], 9), format_map, 'online'),
                    region=cell(row, ['Регион', 'region'], 10) or 'Россия',
                    short_description=cell(row, ['Краткое описание', 'short_description', 'summary'], 11) or '',
                    description=cell(row, ['Подробное описание', 'Описание', 'description', 'details'], 12) or '',
                    is_approved=False,
                )
            return redirect('/admin-panel/?tab=add')

        elif 'title' in request.POST:
            Olympiad.objects.create(
                title=request.POST.get('title'), listing_type=request.POST.get('listing_type') or 'olympiad', organizer=request.POST.get('organizer') or '', link=request.POST.get('link'),
                start_date=request.POST.get('start_date'), end_date=request.POST.get('end_date'),
                target_audience=request.POST.get('target_audience'), grade=request.POST.get('grade') or None,
                stage=request.POST.get('stage'), format=request.POST.get('format'),
                region=request.POST.get('region'), short_description=request.POST.get('short_description') or '', description=request.POST.get('description'),
                cover_image=request.FILES.get('cover_image') if 'cover_image' in request.FILES else None,
                is_approved=False
            )
            return redirect('/admin-panel/?tab=add')

        return redirect('/admin-panel/')

    all_approved = Olympiad.objects.filter(is_approved=True).prefetch_related('comments__user__profile')
    active_olymps = [o for o in all_approved if o.is_active]
    archive_olymps = [o for o in all_approved if not o.is_active]

    user_rows = []
    for profile in Profile.objects.select_related('user').all().order_by('user__username'):
        user_rows.append({
            'profile': profile,
            'last_seen': profile.last_seen,
            'is_online': profile.is_online,
        })

    support_threads_qs = SupportThread.objects.select_related('user', 'user__profile', 'comment').prefetch_related('messages').all()
    thread_rows = []
    for thread in support_threads_qs:
        Profile.objects.get_or_create(user=thread.user)
        thread_rows.append({'thread': thread, 'unread': thread.messages.filter(is_admin=False, read_by_admin=False).count()})

    presets = get_theme_presets()
    edit_theme = normalize_theme(request.GET.get('edit_theme') or get_setting('default_theme', 'telegram-light'))
    selected_preset = active_theme_preset(edit_theme)

    context = {
        'tab': tab,
        'active_olymps': active_olymps,
        'pending_olymps': Olympiad.objects.filter(is_approved=False).prefetch_related('comments__user__profile'),
        'archive_olymps': archive_olymps,
        'pending_comments': Comment.objects.filter(is_approved=False, is_deleted=False),
        'all_profiles': user_rows,
        'olymp_to_edit': olymp_to_edit,
        'custom_theme': custom_theme,
        'banned_words': BannedWord.objects.all().order_by('word'),
        'theme_presets': presets,
        'theme_choices': [(p.key, p.name) for p in presets],
        'default_theme': normalize_theme(get_setting('default_theme', 'telegram-light')),
        'rules_content': get_setting('rules_content', DEFAULT_RULES_CONTENT),
        'cookies_content': get_setting('cookies_content', DEFAULT_COOKIES_CONTENT),
        'support_threads': support_threads_qs,
        'thread_rows': thread_rows,
        'selected_thread': selected_thread,
        'edit_theme': edit_theme,
        'selected_preset': selected_preset,
        'media_files': build_media_inventory() if tab == 'media' else [],
    }

    return render_main(request, 'main/admin.html', context)


def export_excel(request):
    if not (is_site_admin(request.user)):
        return redirect('login_admin')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Олимпиады'
    ws.append(['Номер', 'Тип', 'Название', 'Организатор', 'Начало', 'Конец', 'Ссылка', 'Аудитория', 'Класс', 'Уровень', 'Формат', 'Регион', 'Краткое описание', 'Подробное описание', 'Проверено'])
    for o in Olympiad.objects.all().order_by('public_number', 'id'):
        ws.append([o.public_code, o.get_listing_type_display(), o.title, o.organizer, o.start_date, o.end_date, o.link, o.get_target_audience_display(), o.grade, o.get_stage_display(), o.get_format_display(), o.region, o.short_description, o.description, o.is_approved])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=olympiads.xlsx'
    wb.save(response)
    return response


def excel_instruction(request):
    if not (is_site_admin(request.user)):
        return redirect('login_admin')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Шаблон'
    headers = ['Номер', 'Тип', 'Название', 'Организатор', 'Начало', 'Конец', 'Ссылка', 'Аудитория', 'Класс', 'Уровень', 'Формат', 'Регион', 'Краткое описание', 'Подробное описание']
    ws.append(headers)
    ws.append(['#0000', 'Олимпиада', 'Название события', 'Организатор', '2026-09-01', '2026-10-01', 'https://example.com', 'Школьник', '9', 'Региональный', 'Онлайн', 'Москва', 'Коротко для карточки', 'Полное описание для страницы олимпиады'])
    ws.append(['', 'Грант', 'Название гранта', 'Фонд / компания', '2026-09-15', '2026-12-01', 'https://example.com', 'Студент', '', 'Всероссийский', 'Очно', 'Россия', 'Краткое описание гранта', 'Если номер пустой, сайт выдаст его сам. Если номер уже есть, строка не загрузится повторно.'])
    ws2 = wb.create_sheet('Инструкция')
    ws2.append(['Правила заполнения'])
    ws2.append(['Номер можно оставить пустым. Сайт выдаст #0000, #0001 и далее автоматически.'])
    ws2.append(['Тип: Олимпиада или Грант.'])
    ws2.append(['Аудитория: Школьник или Студент.'])
    ws2.append(['Уровень: Муниципальный, Региональный или Всероссийский.'])
    ws2.append(['Формат: Онлайн или Очно.'])
    ws2.append(['Даты лучше писать в формате ГГГГ-ММ-ДД.'])
    ws2.append(['Краткое описание используется на карточке. Подробное описание открывается на отдельной странице.'])
    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=olympiads_import_instruction.xlsx'
    wb.save(response)
    return response


def settings_view(request):
    if request.method == 'POST':
        selected_theme = normalize_theme(request.POST.get('theme') or request.session.get('theme') or get_setting('default_theme', 'telegram-light'))
        request.session['theme'] = selected_theme
        request.session['visually_impaired'] = 'visually_impaired' in request.POST
        request.session['sound_enabled'] = '1' if 'sound_enabled' in request.POST else '0'
    return redirect(request.META.get('HTTP_REFERER', 'index'))


def logout_admin(request):
    if request.user.is_authenticated:
        Profile.objects.filter(user=request.user).update(last_activity=None)
    logout(request)
    request.session.flush()
    return redirect('index')
